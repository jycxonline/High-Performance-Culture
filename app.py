"""
High Performance Culture Diagnostic Tool
========================================
Streamlit application implementing the full specification:
  · Employee questionnaire
  · Admin dashboard (department comparison + filters)
  · Upload master configuration (with diff preview)
  · Upload response data
  · Generate executive PDF report
"""
from __future__ import annotations
import io
import uuid
import shutil
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import streamlit as st

from hpc.config_loader import load_config, diff_configs, append_change_log, HPCConfig, PILLARS
from hpc.engine import analyze, load_responses, append_submission
from hpc import charts
from hpc.report import build_pdf

# ---------------------------------------------------------------------------
# App configuration
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="High Performance Culture Diagnostic Tool",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

CONFIG_PATH = DATA_DIR / "HPC_Question_Bank_Template.xlsx"
RESPONSE_PATH = DATA_DIR / "HPC_Response_Data_Template.xlsx"
GENERATED_DIR = DATA_DIR / "generated"
GENERATED_DIR.mkdir(exist_ok=True)

NAVY = "#1F3864"
GOLD = "#BF9000"


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
CUSTOM_CSS = f"""
<style>
    .main .block-container {{ padding-top: 1.5rem; padding-bottom: 2rem; max-width: 1300px; }}
    h1, h2, h3 {{ color: {NAVY}; }}
    .stButton>button {{
        background-color: {NAVY}; color: white; border: 0; padding: 0.5rem 1.2rem;
        border-radius: 4px; font-weight: 600;
    }}
    .stButton>button:hover {{ background-color: #2A4A7F; color: white; }}
    .hpc-hero {{
        background: linear-gradient(135deg, {NAVY} 0%, #2A4A7F 100%);
        color: white; padding: 2rem 2.5rem; border-radius: 8px; margin-bottom: 1.5rem;
        border-bottom: 3px solid {GOLD};
    }}
    .hpc-hero h1 {{ color: white; margin: 0; font-size: 2rem; }}
    .hpc-hero p  {{ color: #D9E2F3; margin: 0.5rem 0 0 0; }}
    .hpc-callout {{
        background: #D9E2F3; padding: 1rem 1.2rem; border-left: 4px solid {NAVY};
        border-radius: 4px; margin: 0.8rem 0;
    }}
    .hpc-metric-card {{
        background: white; padding: 1.2rem; border-radius: 6px;
        border: 1px solid #E7E7E7; box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }}
    .band-chip {{
        display: inline-block; padding: 4px 10px; border-radius: 12px;
        font-size: 0.85rem; font-weight: 600; color: white;
    }}
    .band-dysf {{ background: #C00000; }}
    .band-bal  {{ background: #ED7D31; }}
    .band-perf {{ background: #4472C4; }}
    .band-hpc  {{ background: #548235; }}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


def band_chip_html(cls: str) -> str:
    mapping = {
        "Dysfunctional Culture": "band-dysf",
        "Balanced Culture": "band-bal",
        "Performing Culture": "band-perf",
        "High Performance Culture": "band-hpc",
    }
    return f'<span class="band-chip {mapping.get(cls, "band-perf")}">{cls}</span>'


# ---------------------------------------------------------------------------
# Data loading (cached)
# ---------------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def _load_cfg(path: str, mtime: float) -> HPCConfig:
    return load_config(path)


@st.cache_data(show_spinner=False)
def _load_responses(path: str, mtime: float) -> pd.DataFrame:
    return load_responses(path)


def get_config() -> HPCConfig:
    if not CONFIG_PATH.exists():
        st.error(f"Configuration file not found at {CONFIG_PATH}. Please upload one via 'Upload Configuration'.")
        st.stop()
    return _load_cfg(str(CONFIG_PATH), CONFIG_PATH.stat().st_mtime)


def get_responses() -> pd.DataFrame:
    if not RESPONSE_PATH.exists():
        return pd.DataFrame(columns=[
            "Submission ID", "Submission Timestamp", "Department", "Respondent ID",
            "Question ID", "Question Text", "Pillar", "Score"
        ])
    return _load_responses(str(RESPONSE_PATH), RESPONSE_PATH.stat().st_mtime)


def clear_caches():
    st.cache_data.clear()


# ---------------------------------------------------------------------------
# Sidebar navigation
# ---------------------------------------------------------------------------
with st.sidebar:
    st.markdown(f"<h2 style='color:{NAVY};margin:0'>HPC Diagnostic</h2>", unsafe_allow_html=True)
    st.markdown(f"<p style='color:#666;margin:0 0 1rem 0'>Powered by the PERILL framework</p>",
                unsafe_allow_html=True)
    page = st.radio(
        "Navigation",
        ["🏠 Home",
         "📝 Take Questionnaire",
         "📊 Admin Dashboard",
         "📥 Upload Response Data",
         "⚙️ Upload Configuration",
         "📄 Generate Executive Report"],
        label_visibility="collapsed",
    )
    st.markdown("---")
    # Quick stats
    try:
        cfg = get_config()
        st.caption(f"**Active questions:** {len(cfg.active_questions)}")
        st.caption(f"**Active departments:** {len(cfg.active_departments)}")
        st.caption(f"**Tool version:** {cfg.get('tool_version', '1.0')}")
    except Exception:
        pass
    try:
        n_subs = get_responses()["Submission ID"].nunique()
        st.caption(f"**Total submissions:** {n_subs}")
    except Exception:
        pass


# ---------------------------------------------------------------------------
# PAGE — HOME
# ---------------------------------------------------------------------------
def page_home():
    st.markdown(
        f"""<div class="hpc-hero">
            <h1>High Performance Culture Diagnostic Tool</h1>
            <p>Diagnose culture across Purpose · Partnership · Processes · Excellence.
            Built on the PERILL framework.</p>
        </div>""",
        unsafe_allow_html=True,
    )

    cfg = get_config()
    df = get_responses()

    if len(df) == 0:
        st.info("No response data loaded yet. Try 'Take Questionnaire' or 'Upload Response Data'.")
        return

    # Company-wide snapshot
    n_subs = df["Submission ID"].nunique()
    n_depts = df["Department"].nunique()
    company_pillar = df.groupby("Pillar")["Score"].mean().reindex(PILLARS)
    company_overall = float(company_pillar.mean())
    from hpc.engine import classify_score, balance_label
    company_class = classify_score(company_overall, cfg)
    imbalance = float(company_pillar.max() - company_pillar.min())

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total submissions", f"{n_subs:,}")
    c2.metric("Departments", n_depts)
    c3.metric("Company HPC score", f"{company_overall:.2f}", f"{imbalance:.2f} gap")
    c4.markdown(f"**Company classification**<br>{band_chip_html(company_class)}",
                unsafe_allow_html=True)

    st.markdown("### Company-wide pillar scores")
    pillar_df = pd.DataFrame({
        "Pillar": PILLARS,
        "Score": [round(float(company_pillar[p]), 2) for p in PILLARS],
        "Status": [classify_score(float(company_pillar[p]), cfg) for p in PILLARS],
    })
    st.dataframe(pillar_df, use_container_width=True, hide_index=True)

    st.markdown("### How to use this tool")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown(
            "**For employees**\n"
            "1. Go to *Take Questionnaire*\n"
            "2. Select your department\n"
            "3. Answer 40 questions (1–10 scale)\n"
            "4. Submit — your responses are anonymous"
        )
    with col2:
        st.markdown(
            "**For administrators**\n"
            "1. Explore results in *Admin Dashboard*\n"
            "2. Upload new data via *Upload Response Data*\n"
            "3. Adjust questions/departments via *Upload Configuration*\n"
            "4. Download the PDF via *Generate Executive Report*"
        )


# ---------------------------------------------------------------------------
# PAGE — QUESTIONNAIRE
# ---------------------------------------------------------------------------
def page_questionnaire():
    st.markdown(
        f"""<div class="hpc-hero">
            <h1>Take the Questionnaire</h1>
            <p>Your feedback shapes how we build a high performance culture.
            Anonymous, ~5 minutes, 40 questions.</p>
        </div>""", unsafe_allow_html=True,
    )

    cfg = get_config()
    active = cfg.active_questions
    depts = cfg.active_departments

    if not depts:
        st.error("No active departments configured. Please update the configuration file.")
        return

    # Initialize state
    if "quest_step" not in st.session_state:
        st.session_state.quest_step = "start"
    if "quest_answers" not in st.session_state:
        st.session_state.quest_answers = {}

    # ---- Step 1: intro/select department ----
    if st.session_state.quest_step == "start":
        st.markdown(f"""
        <div class="hpc-callout">
            <b>How the scale works:</b> Every question is rated on a <b>{cfg.scale_min}–{cfg.scale_max}</b> scale.<br>
            <b>{cfg.scale_min}</b> = {cfg.scale_label_min}  ·  <b>{cfg.scale_max}</b> = {cfg.scale_label_max}
        </div>
        """, unsafe_allow_html=True)
        dept = st.selectbox("Please select your department", [""] + depts)
        st.caption(f"You will answer {len(active)} questions across four pillars: {', '.join(PILLARS)}.")
        if st.button("Start questionnaire", type="primary", disabled=(dept == "")):
            st.session_state.quest_dept = dept
            st.session_state.quest_step = "answer"
            st.session_state.quest_answers = {}
            st.rerun()
        return

    # ---- Step 2: answer questions ----
    if st.session_state.quest_step == "answer":
        st.markdown(f"**Department:** {st.session_state.quest_dept}")
        answered = len(st.session_state.quest_answers)
        total = len(active)
        st.progress(answered / total, text=f"{answered} / {total} answered")

        # Group by pillar with tabs
        tabs = st.tabs(PILLARS)
        for tab, pillar in zip(tabs, PILLARS):
            with tab:
                pillar_qs = active[active["Pillar"] == pillar]
                for _, row in pillar_qs.iterrows():
                    qid = row["Question ID"]
                    st.markdown(f"**{qid}.** {row['Question Text']}")
                    current = st.session_state.quest_answers.get(qid, None)
                    val = st.slider(
                        f"Rate ({cfg.scale_min} = {cfg.scale_label_min}, {cfg.scale_max} = {cfg.scale_label_max})",
                        min_value=cfg.scale_min, max_value=cfg.scale_max,
                        value=current if current is not None else int((cfg.scale_min + cfg.scale_max) / 2),
                        key=f"q_{qid}", label_visibility="collapsed",
                    )
                    st.session_state.quest_answers[qid] = val
                    st.markdown("---")

        # Bottom bar
        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            if st.button("← Back"):
                st.session_state.quest_step = "start"
                st.rerun()
        with col3:
            can_submit = len(st.session_state.quest_answers) == len(active)
            if st.button("Submit responses ✓", type="primary", disabled=not can_submit):
                # Persist
                sub_id = f"HPC-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
                answers_map = {}
                for _, row in active.iterrows():
                    qid = row["Question ID"]
                    answers_map[qid] = (row["Question Text"], row["Pillar"],
                                        int(st.session_state.quest_answers[qid]))
                _ensure_response_file(cfg)
                append_submission(
                    str(RESPONSE_PATH),
                    submission_id=sub_id,
                    department=st.session_state.quest_dept,
                    respondent_id="" if cfg.anonymous_mode else f"USER-{uuid.uuid4().hex[:6].upper()}",
                    answers=answers_map,
                )
                clear_caches()
                st.session_state.quest_step = "done"
                st.session_state.quest_sub_id = sub_id
                st.rerun()

    # ---- Step 3: thank you ----
    if st.session_state.quest_step == "done":
        st.balloons()
        st.success("✅ Thank you — your responses have been submitted.")
        st.markdown(f"""
        <div class="hpc-callout">
            <b>Submission ID:</b> {st.session_state.quest_sub_id}<br>
            <b>Department:</b> {st.session_state.quest_dept}<br>
            Your responses are anonymous and will be aggregated with others in your department.
        </div>
        """, unsafe_allow_html=True)
        if st.button("Take another"):
            st.session_state.quest_step = "start"
            st.session_state.quest_answers = {}
            st.rerun()


def _ensure_response_file(cfg: HPCConfig):
    """Create an empty response Excel file if one doesn't exist."""
    if RESPONSE_PATH.exists():
        return
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"
    headers = ["Submission ID", "Submission Timestamp", "Department", "Respondent ID",
               "Question ID", "Question Text", "Pillar", "Score"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F3864")
        c.alignment = Alignment(horizontal="center")
    wb.save(RESPONSE_PATH)


# ---------------------------------------------------------------------------
# PAGE — ADMIN DASHBOARD
# ---------------------------------------------------------------------------
def page_dashboard():
    st.markdown(
        f"""<div class="hpc-hero">
            <h1>Admin Dashboard</h1>
            <p>Filter, compare, and drill into culture data across departments.</p>
        </div>""", unsafe_allow_html=True,
    )
    cfg = get_config()
    df = get_responses()
    if len(df) == 0:
        st.warning("No response data available yet. Please collect responses or upload a dataset.")
        return

    # ---- Filters ----
    with st.expander("🔍 Filters", expanded=True):
        c1, c2, c3 = st.columns(3)
        available_depts = sorted(df["Department"].unique())
        with c1:
            selected = st.multiselect("Departments", available_depts, default=available_depts)
        with c2:
            # Date range
            df["_ts"] = pd.to_datetime(df["Submission Timestamp"], errors="coerce")
            min_d = df["_ts"].min()
            max_d = df["_ts"].max()
            if pd.notna(min_d) and pd.notna(max_d):
                dr = st.date_input("Date range", value=(min_d.date(), max_d.date()),
                                   min_value=min_d.date(), max_value=max_d.date())
            else:
                dr = None
        with c3:
            min_resp = st.number_input("Minimum responses per department",
                                       min_value=1, value=cfg.min_responses_dept, step=1)

    # Apply filters
    filtered = df[df["Department"].isin(selected)].copy()
    if dr and len(dr) == 2:
        filtered = filtered[(filtered["_ts"] >= pd.to_datetime(dr[0])) &
                            (filtered["_ts"] <= pd.to_datetime(dr[1]) + pd.Timedelta(days=1))]

    dept_counts = filtered.groupby("Department")["Submission ID"].nunique()
    small_depts = dept_counts[dept_counts < min_resp].index.tolist()
    if small_depts:
        st.warning(f"⚠️ Interpret results with caution. These departments have fewer than "
                   f"{min_resp} responses: {', '.join(small_depts)}")

    if len(filtered) == 0:
        st.error("No data matches the current filters.")
        return

    # ---- Run analysis for company-wide view ----
    analysis = analyze(filtered, "__ALL__", cfg)

    # ---- KPI row ----
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Submissions", f"{analysis.company_n:,}")
    k2.metric("Departments", len(analysis.all_departments))
    k3.metric("Overall HPC score", f"{analysis.company_overall:.2f}")
    k4.markdown(f"**Classification**<br>{band_chip_html(analysis.focus.classification)}",
                unsafe_allow_html=True)

    # ---- Tabs ----
    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        ["📊 Pillar overview", "🕸️ Radar comparison", "🔥 Correlation heatmap",
         "📋 Department table", "💡 Insights"])

    with tab1:
        st.markdown("### Pillar scores by department")
        pv = analysis.all_departments[PILLARS + ["Overall", "N respondents"]].round(2)
        pv = pv.sort_values("Overall", ascending=False)
        st.dataframe(pv, use_container_width=True)

        st.markdown("### Department ranking")
        fig = charts.ranking_bar(analysis.all_departments, analysis.company_overall)
        st.pyplot(fig, use_container_width=True)

    with tab2:
        st.markdown("### Overlay departments on the radar")
        overlay = st.multiselect(
            "Choose up to 5 departments to compare",
            sorted(analysis.all_departments.index),
            default=sorted(analysis.all_departments.index)[:3],
            max_selections=5,
        )
        if overlay:
            dept_means = {d: {p: float(analysis.all_departments.loc[d, p]) for p in PILLARS}
                          for d in overlay}
            fig = charts.radar_chart_multi(dept_means, analysis.company_pillar_means)
            st.pyplot(fig, use_container_width=True)

    with tab3:
        st.markdown("### How the pillars move together")
        fig = charts.correlation_heatmap(analysis.correlation)
        st.pyplot(fig, use_container_width=True)
        st.caption("Strong positive correlations (dark blue) indicate mutually reinforcing pillars. "
                   "Negative correlations (red) suggest tension between pillars.")

    with tab4:
        st.markdown("### Full department table")
        table = analysis.all_departments.copy()
        table["Overall"] = table["Overall"].round(2)
        table["Imbalance"] = table["Imbalance"].round(2)
        # Add classification column
        from hpc.engine import classify_score, apply_imbalance_downgrade
        classes = []
        for d in table.index:
            base = classify_score(table.loc[d, "Overall"], cfg)
            final, dg = apply_imbalance_downgrade(base, table.loc[d, "Imbalance"], cfg)
            classes.append(final + (" ⚠" if dg else ""))
        table["Classification"] = classes
        st.dataframe(table.round(2), use_container_width=True)

    with tab5:
        st.markdown("### Rule-based insights (company-wide view)")
        for ins in analysis.insights:
            st.markdown(f"- **{ins['label']}.** {ins['text']}  *[{ins['rule']}]*")


# ---------------------------------------------------------------------------
# PAGE — UPLOAD RESPONSE DATA
# ---------------------------------------------------------------------------
def page_upload_responses():
    st.markdown(
        f"""<div class="hpc-hero">
            <h1>Upload Response Data</h1>
            <p>Import a response dataset that matches the HPC schema. Supports .xlsx and .csv.</p>
        </div>""", unsafe_allow_html=True,
    )

    st.markdown("### Required schema")
    st.markdown("""
    The file must contain a sheet named **Responses** (or be a single-sheet workbook / CSV)
    with these columns:

    | Column | Description |
    |---|---|
    | Submission ID | Unique per completed questionnaire |
    | Submission Timestamp | ISO date/time |
    | Department | Must match an active department |
    | Respondent ID | Optional |
    | Question ID | Must match the Question Bank |
    | Question Text | Immutable copy at time of submission |
    | Pillar | Purpose / Partnership / Processes / Excellence |
    | Score | Integer 1–10 |
    """)

    up = st.file_uploader("Choose response file", type=["xlsx", "csv"])
    mode = st.radio("Upload mode", ["Replace existing data", "Append to existing data"],
                    horizontal=True)

    if up is not None:
        # Save a temp file
        tmp = DATA_DIR / f"_incoming_{up.name}"
        tmp.write_bytes(up.getvalue())

        try:
            incoming = load_responses(str(tmp))
        except Exception as e:
            st.error(f"Could not read file: {e}")
            tmp.unlink(missing_ok=True)
            return

        st.success(f"✅ File is valid. {len(incoming):,} response rows detected "
                   f"across {incoming['Department'].nunique()} departments.")
        st.dataframe(incoming.head(10), use_container_width=True)

        if st.button("Apply upload", type="primary"):
            if mode == "Replace existing data":
                shutil.copy(tmp, RESPONSE_PATH)
            else:
                existing = get_responses() if RESPONSE_PATH.exists() else pd.DataFrame()
                combined = pd.concat([existing, incoming], ignore_index=True)
                combined.to_excel(RESPONSE_PATH, sheet_name="Responses", index=False)
            tmp.unlink(missing_ok=True)
            clear_caches()
            st.success(f"✅ Response data updated. Refresh to see changes.")
            st.balloons()


# ---------------------------------------------------------------------------
# PAGE — UPLOAD CONFIGURATION
# ---------------------------------------------------------------------------
def page_upload_config():
    st.markdown(
        f"""<div class="hpc-hero">
            <h1>Upload Configuration</h1>
            <p>Edit questions, departments and settings in the master Excel file, then upload here.
            The tool validates the schema and shows you a diff before applying.</p>
        </div>""", unsafe_allow_html=True,
    )

    current = get_config()
    st.markdown("### Current configuration")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Active questions", len(current.active_questions))
    c2.metric("Active departments", len(current.active_departments))
    c3.metric("Tool version", str(current.get("tool_version", "1.0")))
    c4.metric("Schema", current.schema_version)

    with st.expander("Download the current configuration file"):
        with open(CONFIG_PATH, "rb") as f:
            st.download_button("Download HPC_Question_Bank_Template.xlsx", f.read(),
                               file_name="HPC_Question_Bank_Template.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("---")
    st.markdown("### Upload new configuration")
    up = st.file_uploader("Choose master configuration Excel file", type=["xlsx"])

    if up is not None:
        tmp = DATA_DIR / f"_incoming_config_{up.name}"
        tmp.write_bytes(up.getvalue())

        try:
            incoming = load_config(str(tmp))
        except Exception as e:
            st.error(f"❌ Configuration file rejected: {e}")
            tmp.unlink(missing_ok=True)
            return

        st.success(f"✅ Configuration file is valid (schema: {incoming.schema_version}).")
        c1, c2 = st.columns(2)
        c1.metric("Incoming: active questions", len(incoming.active_questions))
        c2.metric("Incoming: active departments", len(incoming.active_departments))

        # Diff preview
        st.markdown("### Diff preview")
        diff = diff_configs(current, incoming)
        any_change = any(len(v) > 0 for v in diff.values())
        if not any_change:
            st.info("No changes detected between the current and incoming configuration.")
        else:
            for category, items in diff.items():
                if items:
                    with st.expander(f"{category} ({len(items)})", expanded=(len(items) < 5)):
                        for item in items:
                            st.markdown(f"- {item}")

        # Change log entry
        st.markdown("### Change log entry (required)")
        c1, c2 = st.columns(2)
        with c1:
            changed_by = st.text_input("Changed by", value=str(incoming.get("last_updated_by", "")))
        with c2:
            change_type = st.selectbox("Change type", [
                "Configuration update", "Question edit", "Question added", "Question retired",
                "Department added", "Department renamed", "Department retired", "Config change", "Other",
            ])
        summary = st.text_area("Summary of change", height=80,
                               placeholder="Describe what changed and why...")

        if st.button("Apply configuration ✓", type="primary",
                     disabled=(not changed_by or not summary)):
            # Backup old config
            backup = DATA_DIR / f"HPC_Question_Bank_Template.backup.{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            shutil.copy(CONFIG_PATH, backup)
            # Apply new
            shutil.copy(tmp, CONFIG_PATH)
            tmp.unlink(missing_ok=True)
            # Append change log
            append_change_log(str(CONFIG_PATH), [{
                "Date": date.today().isoformat(),
                "Changed By": changed_by,
                "Change Type": change_type,
                "Item Affected": "Multiple",
                "Summary of Change": summary,
                "Reason / Approval": "Uploaded via HPC Diagnostic Tool",
            }])
            clear_caches()
            st.success(f"✅ Configuration applied. Previous file backed up to {backup.name}.")
            st.balloons()


# ---------------------------------------------------------------------------
# PAGE — GENERATE REPORT
# ---------------------------------------------------------------------------
def page_generate_report():
    st.markdown(
        f"""<div class="hpc-hero">
            <h1>Generate Executive Report</h1>
            <p>Produce the professional PDF report used for leadership briefings.</p>
        </div>""", unsafe_allow_html=True,
    )

    cfg = get_config()
    df = get_responses()
    if len(df) == 0:
        st.warning("No response data available. Please collect or upload data first.")
        return

    available = ["Company-wide (all departments)"] + sorted(df["Department"].unique())
    focus = st.selectbox("Select the focus department", available)
    prepared_by = st.text_input("Prepared by", value="Organization Development — Culture")
    date_str = st.text_input("Date of analysis", value=datetime.now().strftime("%d %B %Y"))

    if st.button("Generate PDF ✓", type="primary"):
        with st.spinner("Building executive report..."):
            focus_key = "__ALL__" if focus.startswith("Company-wide") else focus
            analysis = analyze(df, focus_key, cfg)
            fn = f"HPC_Diagnostic_Report_{focus.replace(' ', '_').replace('(', '').replace(')', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            out_path = GENERATED_DIR / fn
            build_pdf(analysis, cfg, str(out_path), prepared_by=prepared_by,
                      analysis_date=date_str)

        st.success("✅ Report generated.")

        # KPI summary
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Overall HPC score", f"{analysis.focus.overall:.2f}")
        c2.metric("Company average", f"{analysis.company_overall:.2f}",
                  f"{analysis.focus.overall - analysis.company_overall:+.2f}")
        c3.metric("Pillar gap", f"{analysis.focus.imbalance:.2f}")
        c4.markdown(f"**Classification**<br>{band_chip_html(analysis.focus.classification)}",
                    unsafe_allow_html=True)

        with open(out_path, "rb") as f:
            st.download_button("📄 Download PDF report", f.read(),
                               file_name=fn, mime="application/pdf")


# ---------------------------------------------------------------------------
# Router
# ---------------------------------------------------------------------------
if page == "🏠 Home":
    page_home()
elif page == "📝 Take Questionnaire":
    page_questionnaire()
elif page == "📊 Admin Dashboard":
    page_dashboard()
elif page == "📥 Upload Response Data":
    page_upload_responses()
elif page == "⚙️ Upload Configuration":
    page_upload_config()
elif page == "📄 Generate Executive Report":
    page_generate_report()

# Footer
st.markdown(
    f"""<div style="margin-top:3rem;padding-top:1rem;border-top:1px solid #E7E7E7;
    color:#8C8C8C;font-size:0.85rem;text-align:center">
    High Performance Culture Diagnostic Tool v1.0 · Powered by the PERILL framework ·
    Confidential
    </div>""",
    unsafe_allow_html=True,
)
