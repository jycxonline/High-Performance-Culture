"""
Load and validate the Master Configuration File
(HPC_Question_Bank_Template.xlsx).

The file is treated as the single source of truth for:
  - Question bank
  - Departments
  - Global settings (scale, bands, thresholds)
  - Change log
"""
from __future__ import annotations
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


REQUIRED_SHEETS = ["Question Bank", "Departments", "Config"]
SCHEMA_TAG = "HPC-CONFIG-v1"
PILLARS = ["Purpose", "Partnership", "Processes", "Excellence"]


@dataclass
class HPCConfig:
    """Parsed configuration from the master Excel file."""
    questions: pd.DataFrame
    departments: pd.DataFrame
    settings: dict[str, Any]
    change_log: pd.DataFrame
    schema_version: str = SCHEMA_TAG
    source_path: str | None = None

    # -----------------------------------------------------------
    # Convenience accessors
    # -----------------------------------------------------------
    @property
    def active_questions(self) -> pd.DataFrame:
        df = self.questions.copy()
        return df[df["Active / Inactive"].str.strip().str.lower() == "active"] \
            .sort_values("Display Order").reset_index(drop=True)

    @property
    def active_departments(self) -> list[str]:
        df = self.departments
        actives = df[df["Active / Inactive"].str.strip().str.lower() == "active"]
        return actives.sort_values("Display Order")["Department Name"].tolist()

    def get(self, key: str, default: Any = None) -> Any:
        return self.settings.get(key, default)

    # Scoring thresholds (with safe fallbacks)
    @property
    def band_dysfunctional_max(self) -> float:
        return float(self.get("band_dysfunctional_max", 3.99))

    @property
    def band_balanced_max(self) -> float:
        return float(self.get("band_balanced_max", 5.99))

    @property
    def band_performing_max(self) -> float:
        return float(self.get("band_performing_max", 7.99))

    @property
    def imbalance_moderate_max(self) -> float:
        return float(self.get("imbalance_moderate_max", 2.00))

    @property
    def imbalance_wellbalanced_max(self) -> float:
        return float(self.get("imbalance_wellbalanced_max", 1.00))

    @property
    def min_responses_dept(self) -> int:
        return int(self.get("min_responses_dept", 10))

    @property
    def scale_min(self) -> int:
        return int(self.get("scale_min", 1))

    @property
    def scale_max(self) -> int:
        return int(self.get("scale_max", 10))

    @property
    def scale_label_min(self) -> str:
        return str(self.get("scale_label_min", "Strongly Disagree"))

    @property
    def scale_label_max(self) -> str:
        return str(self.get("scale_label_max", "Strongly Agree"))

    @property
    def anonymous_mode(self) -> bool:
        return str(self.get("anonymous_mode", "TRUE")).strip().upper() == "TRUE"


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------
def load_config(path: str | Path) -> HPCConfig:
    """Load and validate the master config Excel file."""
    path = str(path)
    xl = pd.ExcelFile(path)

    # -- Validate required sheets --
    missing = [s for s in REQUIRED_SHEETS if s not in xl.sheet_names]
    if missing:
        raise ValueError(
            f"Configuration file is invalid. Missing required sheet(s): {', '.join(missing)}."
        )

    # -- Optional schema check --
    if "_SCHEMA" in xl.sheet_names:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb["_SCHEMA"]
            found_tag = None
            for row in ws.iter_rows(values_only=True):
                if row and row[0] == "schema_name":
                    found_tag = row[1]
                    break
            if found_tag and found_tag != SCHEMA_TAG:
                raise ValueError(
                    f"Schema mismatch. Expected '{SCHEMA_TAG}', found '{found_tag}'."
                )
        except KeyError:
            pass

    # -- Question Bank --
    questions = pd.read_excel(path, sheet_name="Question Bank")
    questions = questions.dropna(subset=["Question ID", "Question Text"]).copy()
    questions["Question ID"] = questions["Question ID"].astype(str).str.strip()
    questions["Pillar"] = questions["Pillar"].astype(str).str.strip()
    questions["Active / Inactive"] = questions["Active / Inactive"].fillna("Active").astype(str).str.strip()
    questions["Display Order"] = pd.to_numeric(questions["Display Order"], errors="coerce").fillna(9999)

    # Validate pillars
    bad = questions[~questions["Pillar"].isin(PILLARS)]
    if len(bad):
        raise ValueError(
            f"Question Bank contains invalid Pillar values: {sorted(bad['Pillar'].unique())}. "
            f"Allowed: {PILLARS}"
        )

    # -- Departments --
    departments = pd.read_excel(path, sheet_name="Departments")
    departments = departments.dropna(subset=["Dept ID", "Department Name"]).copy()
    departments["Dept ID"] = departments["Dept ID"].astype(str).str.strip()
    departments["Department Name"] = departments["Department Name"].astype(str).str.strip()
    departments["Active / Inactive"] = departments["Active / Inactive"].fillna("Active").astype(str).str.strip()
    departments["Display Order"] = pd.to_numeric(departments["Display Order"], errors="coerce").fillna(9999)

    # -- Config --
    settings_df = pd.read_excel(path, sheet_name="Config")
    settings_df = settings_df.dropna(subset=["Setting Key"]).copy()
    settings = dict(zip(settings_df["Setting Key"].astype(str).str.strip(),
                        settings_df["Value"]))

    # -- Change Log --
    try:
        change_log = pd.read_excel(path, sheet_name="Change Log")
        change_log = change_log.dropna(how="all")
    except Exception:
        change_log = pd.DataFrame(columns=["Date", "Changed By", "Change Type",
                                           "Item Affected", "Summary of Change",
                                           "Reason / Approval"])

    return HPCConfig(
        questions=questions,
        departments=departments,
        settings=settings,
        change_log=change_log,
        source_path=path,
    )


# ---------------------------------------------------------------------------
# Diff preview — supports the "upload" workflow
# ---------------------------------------------------------------------------
def diff_configs(current: HPCConfig, incoming: HPCConfig) -> dict[str, list[str]]:
    """Return a human-readable diff of two configurations."""
    diff: dict[str, list[str]] = {
        "Departments added": [],
        "Departments retired": [],
        "Departments renamed": [],
        "Questions added": [],
        "Questions retired": [],
        "Questions edited": [],
        "Settings changed": [],
    }

    # Departments
    cur_d = current.departments.set_index("Dept ID")
    inc_d = incoming.departments.set_index("Dept ID")

    added_ids = set(inc_d.index) - set(cur_d.index)
    removed_ids = set(cur_d.index) - set(inc_d.index)
    common_ids = set(cur_d.index) & set(inc_d.index)

    for did in sorted(added_ids):
        diff["Departments added"].append(f"{did} — {inc_d.loc[did, 'Department Name']}")
    for did in sorted(removed_ids):
        diff["Departments retired"].append(f"{did} — {cur_d.loc[did, 'Department Name']}")
    for did in sorted(common_ids):
        old = str(cur_d.loc[did, "Department Name"])
        new = str(inc_d.loc[did, "Department Name"])
        if old != new:
            diff["Departments renamed"].append(f"{did}: {old} → {new}")
        old_status = str(cur_d.loc[did, "Active / Inactive"]).strip().lower()
        new_status = str(inc_d.loc[did, "Active / Inactive"]).strip().lower()
        if old_status != new_status:
            diff["Departments retired" if new_status == "inactive" else "Departments added"] \
                .append(f"{did} — {inc_d.loc[did, 'Department Name']} (status: {old_status} → {new_status})")

    # Questions
    cur_q = current.questions.set_index("Question ID")
    inc_q = incoming.questions.set_index("Question ID")

    added_q = set(inc_q.index) - set(cur_q.index)
    removed_q = set(cur_q.index) - set(inc_q.index)
    common_q = set(cur_q.index) & set(inc_q.index)

    for qid in sorted(added_q):
        diff["Questions added"].append(f"{qid} ({inc_q.loc[qid, 'Pillar']}) — {inc_q.loc[qid, 'Question Text'][:80]}")
    for qid in sorted(removed_q):
        diff["Questions retired"].append(f"{qid} — {cur_q.loc[qid, 'Question Text'][:80]}")
    for qid in sorted(common_q):
        old = str(cur_q.loc[qid, "Question Text"]).strip()
        new = str(inc_q.loc[qid, "Question Text"]).strip()
        if old != new:
            diff["Questions edited"].append(f"{qid}: wording changed")
        old_s = str(cur_q.loc[qid, "Active / Inactive"]).strip().lower()
        new_s = str(inc_q.loc[qid, "Active / Inactive"]).strip().lower()
        if old_s != new_s:
            diff["Questions retired" if new_s == "inactive" else "Questions added"] \
                .append(f"{qid} — status: {old_s} → {new_s}")

    # Settings
    for key, new_val in incoming.settings.items():
        old_val = current.settings.get(key)
        if str(old_val) != str(new_val):
            diff["Settings changed"].append(f"{key}: {old_val} → {new_val}")

    return diff


# ---------------------------------------------------------------------------
# Change log helper
# ---------------------------------------------------------------------------
def append_change_log(path: str | Path, entries: list[dict]) -> None:
    """Append rows to the Change Log sheet in-place."""
    wb = load_workbook(path)
    if "Change Log" not in wb.sheetnames:
        return
    ws = wb["Change Log"]
    # Find first empty row
    start = ws.max_row + 1
    for e in entries:
        ws.cell(row=start, column=1, value=e.get("Date", datetime.now().date().isoformat()))
        ws.cell(row=start, column=2, value=e.get("Changed By", ""))
        ws.cell(row=start, column=3, value=e.get("Change Type", ""))
        ws.cell(row=start, column=4, value=e.get("Item Affected", ""))
        ws.cell(row=start, column=5, value=e.get("Summary of Change", ""))
        ws.cell(row=start, column=6, value=e.get("Reason / Approval", ""))
        start += 1
    wb.save(path)
