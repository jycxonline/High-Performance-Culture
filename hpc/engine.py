"""
Analysis engine for the High Performance Culture Diagnostic Tool.

Turns a response dataset + configuration into calculated metrics,
classifications, insights and recommendations. Deterministic and rule-based —
no AI narrative is generated here.
"""
from __future__ import annotations
from dataclasses import dataclass, field
from typing import Any

import numpy as np
import pandas as pd

from .config_loader import HPCConfig, PILLARS


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------
@dataclass
class PillarResult:
    pillar: str
    mean: float
    status: str
    gap_vs_company: float
    interpretation: str
    std: float = 0.0


@dataclass
class DepartmentResult:
    department: str
    n_respondents: int
    pillar_means: dict[str, float]
    overall: float
    imbalance: float
    balance_label: str
    classification: str
    downgraded: bool
    raw_classification: str
    pillar_results: list[PillarResult] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class AnalysisResult:
    focus: DepartmentResult
    company_pillar_means: dict[str, float]
    company_overall: float
    company_n: int
    all_departments: pd.DataFrame        # per-department pillar means + overall
    imbalance_by_dept: pd.Series
    correlation: pd.DataFrame            # pillar × pillar
    insights: list[dict[str, str]]       # rule-based insights
    recommendations: list[dict[str, Any]]
    focus_dept_name: str


# ---------------------------------------------------------------------------
# Scoring & classification
# ---------------------------------------------------------------------------
def classify_score(score: float, cfg: HPCConfig) -> str:
    if score >= cfg.band_performing_max + 0.001:
        return "High Performance Culture"
    if score >= cfg.band_balanced_max + 0.001:
        return "Performing Culture"
    if score >= cfg.band_dysfunctional_max + 0.001:
        return "Balanced Culture"
    return "Dysfunctional Culture"


def balance_label(gap: float, cfg: HPCConfig) -> str:
    if gap <= cfg.imbalance_wellbalanced_max:
        return "Well balanced"
    if gap <= cfg.imbalance_moderate_max:
        return "Moderate imbalance"
    return "Significant imbalance"


def apply_imbalance_downgrade(base: str, gap: float, cfg: HPCConfig) -> tuple[str, bool]:
    """Downgrade classification by one band if gap > moderate max."""
    if gap > cfg.imbalance_moderate_max:
        order = ["Dysfunctional Culture", "Balanced Culture",
                 "Performing Culture", "High Performance Culture"]
        idx = order.index(base)
        if idx > 0:
            return order[idx - 1], True
    return base, False


# ---------------------------------------------------------------------------
# Interpretations (per pillar)
# ---------------------------------------------------------------------------
PILLAR_INTERPRETATIONS = {
    "Purpose": {
        "strong":  "Strategic clarity is present; leadership is connecting work to outcomes.",
        "middle":  "Direction is understood, but line-of-sight to outcomes can be sharpened.",
        "weak":    "Weak strategic clarity — teams are unsure why they exist or what to prioritize.",
    },
    "Partnership": {
        "strong":  "Leadership visibility, trust and collaboration are a genuine strength.",
        "middle":  "Relationships and communication are adequate but not distinctive.",
        "weak":    "Leadership visibility and cross-functional trust need active rebuilding.",
    },
    "Processes": {
        "strong":  "Systems, tools and decision-making enable performance rather than obstruct it.",
        "middle":  "Processes are functional but generate friction under pressure.",
        "weak":    "Primary drag on performance — systems, tools and decisions create friction.",
    },
    "Excellence": {
        "strong":  "Learning culture and growth mindset are the differentiator.",
        "middle":  "Learning and improvement are present but not systematic.",
        "weak":    "Weak learning muscle — improvement is ad hoc rather than a habit.",
    },
}


def _interp(pillar: str, score: float) -> str:
    key = "weak" if score < 5.5 else ("strong" if score >= 7.0 else "middle")
    return PILLAR_INTERPRETATIONS[pillar][key]


# ---------------------------------------------------------------------------
# Main analysis routine
# ---------------------------------------------------------------------------
def analyze(responses: pd.DataFrame,
            focus_dept: str,
            cfg: HPCConfig) -> AnalysisResult:
    """
    Analyze responses and return an AnalysisResult centered on `focus_dept`.
    Set focus_dept="__ALL__" for a company-wide report.
    """
    # Basic validation
    required = {"Department", "Pillar", "Score", "Submission ID"}
    missing = required - set(responses.columns)
    if missing:
        raise ValueError(f"Response data missing columns: {missing}")

    df = responses.copy()
    df["Score"] = pd.to_numeric(df["Score"], errors="coerce")
    df = df.dropna(subset=["Score"])

    # Per-department pillar means
    dept_pillar = (df.groupby(["Department", "Pillar"])["Score"]
                     .mean().unstack()
                     .reindex(columns=PILLARS))

    # Company-wide pillar means
    company_pillar = df.groupby("Pillar")["Score"].mean().reindex(PILLARS)
    company_overall = float(company_pillar.mean())
    company_n = int(df["Submission ID"].nunique())

    # Overall per department
    dept_overall = dept_pillar.mean(axis=1)
    imbalance_by_dept = (dept_pillar.max(axis=1) - dept_pillar.min(axis=1)).sort_values()

    # All-departments summary table
    all_depts = dept_pillar.copy()
    all_depts["Overall"] = dept_overall
    all_depts["Imbalance"] = dept_pillar.max(axis=1) - dept_pillar.min(axis=1)
    all_depts["N respondents"] = df.groupby("Department")["Submission ID"].nunique()

    # Correlation across submissions × pillars
    sub_pillar = (df.groupby(["Submission ID", "Pillar"])["Score"]
                    .mean().unstack().reindex(columns=PILLARS))
    correlation = sub_pillar.corr()

    # -----------------------------------------------------------
    # Focus department
    # -----------------------------------------------------------
    if focus_dept == "__ALL__":
        focus_pillar_series = company_pillar.copy()
        focus_overall = company_overall
        n_focus = company_n
        focus_name = "Company-wide"
        focus_std = df.groupby("Pillar")["Score"].std().reindex(PILLARS)
    else:
        if focus_dept not in dept_pillar.index:
            raise ValueError(f"Department '{focus_dept}' not present in response data.")
        focus_pillar_series = dept_pillar.loc[focus_dept]
        focus_overall = float(dept_overall.loc[focus_dept])
        n_focus = int(df[df.Department == focus_dept]["Submission ID"].nunique())
        focus_name = focus_dept
        focus_std = (df[df.Department == focus_dept]
                     .groupby("Pillar")["Score"].std()
                     .reindex(PILLARS))

    focus_pillar_means = {p: float(focus_pillar_series[p]) for p in PILLARS}
    focus_imbalance = float(focus_pillar_series.max() - focus_pillar_series.min())
    raw_class = classify_score(focus_overall, cfg)
    final_class, downgraded = apply_imbalance_downgrade(raw_class, focus_imbalance, cfg)
    bal_label = balance_label(focus_imbalance, cfg)

    # Pillar results
    pillar_results = []
    for p in PILLARS:
        mean = focus_pillar_means[p]
        std = float(focus_std[p]) if not pd.isna(focus_std[p]) else 0.0
        gap = mean - float(company_pillar[p])
        pillar_results.append(PillarResult(
            pillar=p,
            mean=mean,
            status=classify_score(mean, cfg),
            gap_vs_company=gap,
            interpretation=_interp(p, mean),
            std=std,
        ))

    warnings = []
    if n_focus < cfg.min_responses_dept and focus_dept != "__ALL__":
        warnings.append(
            f"Interpret results with caution. This department has fewer than the recommended "
            f"number of responses ({n_focus} < {cfg.min_responses_dept})."
        )

    focus_result = DepartmentResult(
        department=focus_name,
        n_respondents=n_focus,
        pillar_means=focus_pillar_means,
        overall=focus_overall,
        imbalance=focus_imbalance,
        balance_label=bal_label,
        classification=final_class,
        downgraded=downgraded,
        raw_classification=raw_class,
        pillar_results=pillar_results,
        warnings=warnings,
    )

    # -----------------------------------------------------------
    # Rule-based insights
    # -----------------------------------------------------------
    insights = _generate_insights(focus_result, dept_pillar, dept_overall,
                                  imbalance_by_dept, company_pillar, cfg)

    # Recommendations
    recommendations = _generate_recommendations(focus_result, company_pillar)

    return AnalysisResult(
        focus=focus_result,
        company_pillar_means={p: float(company_pillar[p]) for p in PILLARS},
        company_overall=company_overall,
        company_n=company_n,
        all_departments=all_depts.round(3),
        imbalance_by_dept=imbalance_by_dept,
        correlation=correlation.round(3),
        insights=insights,
        recommendations=recommendations,
        focus_dept_name=focus_name,
    )


# ---------------------------------------------------------------------------
# Insight rules
# ---------------------------------------------------------------------------
def _generate_insights(focus: DepartmentResult,
                       dept_pillar: pd.DataFrame,
                       dept_overall: pd.Series,
                       imbalance_by_dept: pd.Series,
                       company_pillar: pd.Series,
                       cfg: HPCConfig) -> list[dict[str, str]]:
    insights = []
    means = focus.pillar_means
    strong = max(means, key=means.get)
    weak = min(means, key=means.get)
    gap_v_co = {p: means[p] - float(company_pillar[p]) for p in PILLARS}
    largest_gap_pillar = min(gap_v_co, key=gap_v_co.get)

    insights.append({
        "label": "Strongest pillar",
        "text": f"{strong} ({means[strong]:.2f}) — {gap_v_co[strong]:+.2f} vs. company average.",
        "rule": "R-01"
    })
    insights.append({
        "label": "Weakest pillar",
        "text": f"{weak} ({means[weak]:.2f}) — {gap_v_co[weak]:+.2f} vs. company average.",
        "rule": "R-02"
    })
    insights.append({
        "label": "Largest gap vs. company",
        "text": f"{largest_gap_pillar} shows the widest divergence at {gap_v_co[largest_gap_pillar]:+.2f} points.",
        "rule": "R-03"
    })

    # Imbalance flag
    if focus.imbalance > cfg.imbalance_moderate_max:
        insights.append({
            "label": "Imbalance flag",
            "text": (f"Pillar gap of {focus.imbalance:.2f} exceeds the {cfg.imbalance_moderate_max:.2f} threshold. "
                     f"Classification automatically downgraded to {focus.classification}."),
            "rule": "R-04 (triggered)"
        })
    elif focus.imbalance > cfg.imbalance_wellbalanced_max:
        insights.append({
            "label": "Imbalance watch",
            "text": (f"Pillar gap of {focus.imbalance:.2f} sits in the moderate range. "
                     f"Not yet triggering a downgrade, but flagged as a watch item."),
            "rule": "R-04 (watch)"
        })
    else:
        insights.append({
            "label": "Balance",
            "text": f"Pillar gap of {focus.imbalance:.2f} — well balanced.",
            "rule": "R-04"
        })

    # Cross-department context
    if len(dept_pillar) > 1:
        most_balanced = imbalance_by_dept.index[0]
        most_imbalanced = imbalance_by_dept.index[-1]
        insights.append({
            "label": "Most balanced department",
            "text": f"{most_balanced} — pillar gap of {imbalance_by_dept.iloc[0]:.2f}.",
            "rule": "R-05"
        })
        insights.append({
            "label": "Most imbalanced department",
            "text": f"{most_imbalanced} — pillar gap of {imbalance_by_dept.iloc[-1]:.2f}.",
            "rule": "R-06"
        })

    # High-performer confirmation
    if focus.classification == "High Performance Culture":
        insights.append({
            "label": "High-performing confirmation",
            "text": "All pillars above 8.00 and well balanced. Codify and protect what is working.",
            "rule": "R-07"
        })

    return insights


# ---------------------------------------------------------------------------
# Recommendation library
# ---------------------------------------------------------------------------
ACTION_LIBRARY = {
    "Purpose": [
        ("Refresh the strategy cascade — clarify the top 3 priorities and reinforce them at every all-hands.",
         "High", "Rebuilds line-of-sight; lifts Purpose score by 0.5–0.8 within two waves.",
         "Head of Department + Strategy Lead", "0–60 days"),
        ("Run customer immersion sessions for the team — direct exposure to end-user pain.",
         "Medium", "Strengthens customer focus and shared purpose.",
         "Head of Department + Customer Experience Lead", "30–90 days"),
    ],
    "Partnership": [
        ("Introduce a structured joint-planning cadence with top-5 business stakeholders.",
         "High", "Rebuilds trust; reduces reactive workload; lifts collaboration score by 0.5+.",
         "Head of Department + Business Unit heads", "0–60 days"),
        ("Run 'Leader Listening' sessions; convert insights into a public 30-60-90 action plan.",
         "High", "Signals responsiveness; raises Leadership sub-score.",
         "Head of Department + Leadership team", "0–45 days"),
        ("Deliver a psychological-safety intervention for the team (workshop + micro-behaviours).",
         "Medium", "Reduces conflict avoidance; lifts trust and speaking-up scores.",
         "P&C Business Partner", "30–90 days"),
    ],
    "Processes": [
        ("Launch a 60-day rescue programme — target top-3 workflow bottlenecks, tool consolidation and decision-rights redesign.",
         "Critical", "Move Processes score by 0.6–1.0 within two waves; reduce escalations by ~30%.",
         "Head of Department + COO sponsor", "0–90 days"),
        ("Rationalize governance forums — cut redundant meetings by 30% and publish decision rights.",
         "High", "Improves decision speed and role clarity.",
         "Head of Department + Chief of Staff", "30–90 days"),
    ],
    "Excellence": [
        ("Publish a capability roadmap and safeguard L&D budget in the next planning cycle.",
         "High", "Preserves the department's differentiator; prevents erosion under cost pressure.",
         "Head of Department + P&C Business Partner", "30–90 days"),
        ("Introduce a monthly feedback ritual (peer + upward + downward) with a simple template.",
         "Medium", "Builds feedback culture; lifts Excellence sub-scores.",
         "Head of Department + Leadership team", "0–60 days"),
        ("Run innovation sprints on 2 real business challenges each quarter.",
         "Medium", "Demonstrates growth mindset and experimentation.",
         "Head of Department + Innovation champion", "45–120 days"),
    ],
}


def _generate_recommendations(focus: DepartmentResult,
                              company_pillar: pd.Series) -> list[dict[str, Any]]:
    """
    Build a prioritized set of recommendations for the focus department.
    Ordered by pillar weakness, with the weakest pillar getting the highest priority.
    """
    # Order pillars by mean ascending (weakest first)
    ordered = sorted(focus.pillar_means.items(), key=lambda x: x[1])

    recs: list[dict[str, Any]] = []
    for i, (pillar, mean) in enumerate(ordered):
        # Take 1-2 actions from the library per pillar
        actions = ACTION_LIBRARY.get(pillar, [])
        n = 2 if i < 2 else 1  # more for the weakest two
        for action, priority, impact, owner, timeline in actions[:n]:
            # Escalate priority for weakest pillar
            if i == 0 and priority == "High":
                priority = "Critical" if pillar == "Processes" or mean < 4.0 else "High"
            recs.append({
                "Action": action,
                "Pillar": pillar,
                "Priority": priority,
                "Expected Impact": impact,
                "Owner": owner,
                "Timeline": timeline,
            })

    # Always end with a re-measurement recommendation
    recs.append({
        "Action": "Re-run the diagnostic in 6 months; publicly compare movement on the weakest pillars.",
        "Pillar": "All pillars",
        "Priority": "Medium",
        "Expected Impact": "Creates accountability; validates intervention ROI; reinforces culture-as-system message.",
        "Owner": "OD / Culture team",
        "Timeline": "6 months",
    })

    return recs


# ---------------------------------------------------------------------------
# Response ingestion helpers
# ---------------------------------------------------------------------------
def load_responses(path: str) -> pd.DataFrame:
    """Load a response file (accepts the schema of HPC_Response_Data_Template.xlsx)."""
    if str(path).lower().endswith(".csv"):
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path, sheet_name="Responses" if "Responses" in pd.ExcelFile(path).sheet_names else 0)

    required = {"Submission ID", "Department", "Pillar", "Score"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Response file is missing required columns: {missing}")
    return df


def append_submission(path: str,
                      submission_id: str,
                      department: str,
                      respondent_id: str,
                      answers: dict[str, tuple[str, str, int]]) -> None:
    """
    Append a completed questionnaire to the response Excel file.
    `answers` is a dict mapping Question ID → (Question Text, Pillar, Score).
    """
    from openpyxl import load_workbook
    from datetime import datetime as _dt

    wb = load_workbook(path)
    if "Responses" not in wb.sheetnames:
        raise ValueError("Response file has no 'Responses' sheet.")
    ws = wb["Responses"]
    row = ws.max_row + 1
    ts = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
    for qid, (qtext, pillar, score) in answers.items():
        ws.cell(row=row, column=1, value=submission_id)
        ws.cell(row=row, column=2, value=ts)
        ws.cell(row=row, column=3, value=department)
        ws.cell(row=row, column=4, value=respondent_id or "")
        ws.cell(row=row, column=5, value=qid)
        ws.cell(row=row, column=6, value=qtext)
        ws.cell(row=row, column=7, value=pillar)
        ws.cell(row=row, column=8, value=int(score))
        row += 1
    wb.save(path)
